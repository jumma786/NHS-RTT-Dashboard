"""
Enhanced NHS RTT Dashboard — addresses all 16 feedback points.

Sheets:
  1. Dashboard       – 5 charts, dynamic KPIs, clean aligned layout
  2. Data            – Excel Table (slicer-ready), conditional formatting
  3. National_Trend  – monthly national summary
  4. ICB_Summary     – latest-month ranked ICBs
  5. Specialty_Top5  – top 5 specialties + Other (pie-ready)
  6. Top5_Bottom5    – best/worst 5 ICBs
  7. YoY_Change      – year-on-year comparison
  8. Lookups         – dropdown lists
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
NAT_CSV = BASE / "data" / "processed" / "rtt_national.csv"
ICB_CSV = BASE / "data" / "processed" / "rtt_icb.csv"
OUT_PATH = BASE / "dashboard" / "NHS_RTT_Dashboard.xlsx"

# NHS Identity colours
NHS_BLUE = "005EB8"
NHS_DARK = "003087"
NHS_LIGHT = "41B6E6"
NHS_AQUA = "00A9CE"
NHS_GREEN = "009639"
NHS_RED = "DA291C"
NHS_AMBER = "ED8B00"
NHS_GREY = "768692"
NHS_PALE = "E8EDEE"
NHS_BG = "F0F4F5"
WHITE = "FFFFFF"
BLACK = "231F20"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NHS_BLUE, end_color=NHS_BLUE, fill_type="solid")
DARK_FILL = PatternFill(start_color=NHS_DARK, end_color=NHS_DARK, fill_type="solid")
PALE_FILL = PatternFill(start_color=NHS_PALE, end_color=NHS_PALE, fill_type="solid")
BG_FILL = PatternFill(start_color=NHS_BG, end_color=NHS_BG, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color=NHS_GREY),
    right=Side(style="thin", color=NHS_GREY),
    top=Side(style="thin", color=NHS_GREY),
    bottom=Side(style="thin", color=NHS_GREY),
)
NO_BORDER = Border()

# Columns we keep for the ICB data sheet
ICB_KEEP = [
    "Period", "ICB Code", "ICB Name",
    "Treatment Function Code", "Treatment Function",
    "Total number of incomplete pathways",
    "Total within 18 weeks", "% within 18 weeks",
    "Average (median) waiting time (in weeks)",
    "92nd percentile waiting time (in weeks)",
    "Total 52 plus weeks", "Total 65 plus weeks", "Total 78 plus weeks",
]

# Columns we keep for the national data
NAT_KEEP = [
    "Period", "Treatment Function Code", "Treatment Function",
    "Total number of incomplete pathways",
    "Total within 18 weeks", "% within 18 weeks",
    "Average (median) waiting time (in weeks)",
    "92nd percentile waiting time (in weeks)",
    "Total 52 plus weeks", "Total 65 plus weeks", "Total 78 plus weeks",
]

# Short aliases for ICB data sheet columns (readable headers)
SHORT_HEADERS = [
    "Period", "ICB_Code", "ICB_Name",
    "TF_Code", "Treatment_Function",
    "Incomplete_Pathways",
    "Within_18wks", "Pct_Within_18wks",
    "Median_Wait_Wks",
    "P92_Wait_Wks",
    "Waiters_52plus", "Waiters_65plus", "Waiters_78plus",
]

# Custom number format: 1,234,567 → 1.2M / 12,345 → 12.3K
NUM_K = '[>=1000000]#,##0.0,,"M";[>=1000]#,##0.0,"K";#,##0'
NUM_COMMA = "#,##0"
PCT_FMT = "0.0%"
DEC_FMT = "0.0"

CHART_TITLE_FONT_SIZE = 1100  # in hundredths of a point = 11pt


# ── Utilities ─────────────────────────────────────────────────────────

def style_header_row(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, max_width=28):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[letter].width = min(length + 3, max_width)


def write_df(ws, df, pct_cols=None, start_row=1):
    pct_cols = pct_cols or []
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx > start_row:
                cell.alignment = Alignment(horizontal="center")
                col_name = df.columns[c_idx - 1]
                if col_name in pct_cols and isinstance(value, (int, float)):
                    cell.number_format = PCT_FMT
                elif isinstance(value, (int, float)):
                    cell.number_format = NUM_COMMA
    style_header_row(ws, len(df.columns), row=start_row)
    auto_width(ws)


def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, fmt=None, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    return cell


def apply_bg(ws, max_row, max_col, fill):
    """Paint a background fill across unused cells."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill == PatternFill() or cell.fill is None:
                cell.fill = fill


def find_col(ws, name):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == name:
            return c
    return None


def safe_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def fmt_k(n):
    """Format number as K/M string for display."""
    n = safe_int(n)
    if abs(n) >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if abs(n) >= 1_000:
        return f"{n / 1_000:.1f}K"
    return f"{n:,}"


# ── Data preparation ──────────────────────────────────────────────────


def build_national_trend(nat_df):
    nat = nat_df[nat_df["Treatment Function"] == "Total"].copy()
    return nat[NAT_KEEP].sort_values("Period").reset_index(drop=True)


def build_icb_summary(icb_df, latest):
    icb = icb_df[
        (icb_df["Period"] == latest) & (icb_df["Treatment Function"] == "Total")
    ].copy()
    icb["% within 18 weeks"] = pd.to_numeric(icb["% within 18 weeks"], errors="coerce")
    return icb[ICB_KEEP].sort_values("% within 18 weeks", ascending=False).reset_index(drop=True)


def build_specialty_top5(nat_df, latest):
    """Top 5 specialties + 'Other' for clean pie chart. Excludes Total row (#16)."""
    spec = nat_df[
        (nat_df["Period"] == latest) & (nat_df["Treatment Function"] != "Total")
    ].copy()
    spec["Total number of incomplete pathways"] = pd.to_numeric(
        spec["Total number of incomplete pathways"], errors="coerce"
    )
    spec = spec.sort_values("Total number of incomplete pathways", ascending=False)
    top5 = spec.head(5)[["Treatment Function", "Total number of incomplete pathways"]].copy()
    other_sum = spec.iloc[5:]["Total number of incomplete pathways"].sum()
    other_row = pd.DataFrame([{
        "Treatment Function": "Other Specialties",
        "Total number of incomplete pathways": other_sum,
    }])
    result = pd.concat([top5, other_row], ignore_index=True)
    result.columns = ["Specialty", "Incomplete Pathways"]
    return result


def build_top5_bottom5(icb_df, latest):
    icb = icb_df[
        (icb_df["Period"] == latest) & (icb_df["Treatment Function"] == "Total")
    ].copy()
    icb["% within 18 weeks"] = pd.to_numeric(icb["% within 18 weeks"], errors="coerce")
    icb["Total 52 plus weeks"] = pd.to_numeric(icb["Total 52 plus weeks"], errors="coerce")

    # Shorten ICB names for readability (#13)
    icb["ICB Short"] = icb["ICB Name"].str.replace(
        r"^NHS\s+", "", regex=True
    ).str.replace(
        r"\s+INTEGRATED CARE BOARD$", "", regex=True
    )

    best = icb.nlargest(5, "% within 18 weeks")[[
        "ICB Short", "% within 18 weeks", "Total number of incomplete pathways",
        "Average (median) waiting time (in weeks)", "Total 52 plus weeks",
    ]].reset_index(drop=True)
    best.columns = ["ICB", "% Within 18 Wks", "Pathways", "Median Wait", "52+ Waiters"]

    worst = icb.nsmallest(5, "% within 18 weeks")[[
        "ICB Short", "% within 18 weeks", "Total number of incomplete pathways",
        "Average (median) waiting time (in weeks)", "Total 52 plus weeks",
    ]].reset_index(drop=True)
    worst.columns = ["ICB", "% Within 18 Wks", "Pathways", "Median Wait", "52+ Waiters"]

    return best, worst


def build_yoy_change(nat_df):
    nat = nat_df[nat_df["Treatment Function"] == "Total"].copy()
    nat = nat.sort_values("Period")
    cols_numeric = [
        "Total number of incomplete pathways", "% within 18 weeks",
        "Average (median) waiting time (in weeks)", "Total 52 plus weeks",
    ]
    for c in cols_numeric:
        nat[c] = pd.to_numeric(nat[c], errors="coerce")
    nat["Year"] = nat["Period"].str[:4].astype(int)
    nat["Month"] = nat["Period"].str[5:].astype(int)
    result = []
    for _, row in nat.iterrows():
        prev_period = f"{row['Year']-1}-{row['Month']:02d}"
        prev = nat[nat["Period"] == prev_period]
        if prev.empty:
            continue
        p = prev.iloc[0]
        result.append({
            "Period": row["Period"],
            "Pathways": int(row["Total number of incomplete pathways"]),
            "Pathways YoY": int(row["Total number of incomplete pathways"]) - int(p["Total number of incomplete pathways"]),
            "% Within 18 Wks": row["% within 18 weeks"],
            "% Within 18 Wks YoY": row["% within 18 weeks"] - p["% within 18 weeks"],
            "Median Wait": row["Average (median) waiting time (in weeks)"],
            "Median Wait YoY": row["Average (median) waiting time (in weeks)"] - p["Average (median) waiting time (in weeks)"],
            "52+ Waiters": int(row["Total 52 plus weeks"]),
            "52+ Waiters YoY": int(row["Total 52 plus weeks"]) - int(p["Total 52 plus weeks"]),
        })
    return pd.DataFrame(result)


# ── Chart builders ────────────────────────────────────────────────────

CHART_COLORS = [NHS_BLUE, NHS_DARK, NHS_LIGHT, NHS_AQUA, NHS_GREEN, NHS_AMBER]


def add_line_chart(ws, ws_data, title, y_col, nrows, anchor,
                   width=22, height=13, color=NHS_BLUE, y_fmt=None):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height

    col_idx = find_col(ws_data, y_col)
    if not col_idx:
        return

    data = Reference(ws_data, min_col=col_idx, min_row=1, max_row=nrows + 1)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=nrows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    s = chart.series[0]
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = 25000
    s.smooth = False

    if y_fmt:
        chart.y_axis.numFmt = y_fmt
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.numFmt = "General"
    chart.legend = None
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    ws.add_chart(chart, anchor)


def add_dual_line_chart(ws, ws_data, title, col1, col2, nrows, anchor,
                        width=22, height=13):
    """Two-series line chart for comparing two metrics over time."""
    c1 = find_col(ws_data, col1)
    c2 = find_col(ws_data, col2)
    if not c1 or not c2:
        return

    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height

    d1 = Reference(ws_data, min_col=c1, min_row=1, max_row=nrows + 1)
    d2 = Reference(ws_data, min_col=c2, min_row=1, max_row=nrows + 1)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=nrows + 1)

    chart.add_data(d1, titles_from_data=True)
    chart.add_data(d2, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = NHS_BLUE
    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[1].graphicalProperties.line.solidFill = NHS_RED
    chart.series[1].graphicalProperties.line.width = 25000

    chart.x_axis.tickLblPos = "low"
    ws.add_chart(chart, anchor)


def add_bar_chart(ws, ws_data, title, name_col, val_col, nrows, anchor,
                  width=22, height=13, horizontal=True, color=NHS_BLUE,
                  y_fmt=None):
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height

    c_name = find_col(ws_data, name_col)
    c_val = find_col(ws_data, val_col)
    if not c_name or not c_val:
        return

    data = Reference(ws_data, min_col=c_val, min_row=1, max_row=nrows + 1)
    cats = Reference(ws_data, min_col=c_name, min_row=2, max_row=nrows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    s = chart.series[0]
    s.graphicalProperties.solidFill = color

    if y_fmt:
        chart.x_axis.numFmt = y_fmt if horizontal else None
        chart.y_axis.numFmt = None if horizontal else y_fmt

    chart.legend = None
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, ws_data, title, name_col, val_col, nrows, anchor,
                  width=22, height=13):
    chart = PieChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height

    c_name = find_col(ws_data, name_col)
    c_val = find_col(ws_data, val_col)
    if not c_name or not c_val:
        return

    data = Reference(ws_data, min_col=c_val, min_row=1, max_row=nrows + 1)
    cats = Reference(ws_data, min_col=c_name, min_row=2, max_row=nrows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    for i, clr in enumerate(CHART_COLORS[:nrows]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = clr
        chart.series[0].data_points.append(pt)

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showPercent = True
    chart.series[0].dLbls.showCatName = True
    chart.series[0].dLbls.showVal = False

    ws.add_chart(chart, anchor)


# ── Dashboard sheet ───────────────────────────────────────────────────


def build_dashboard(wb, nat_trend, icb_summary, spec_top5, best5, worst5):
    ws = wb.create_sheet("Dashboard")

    # (#1, #2) Background colour — fill entire visible area
    for r in range(1, 95):
        for c in range(1, 22):
            ws.cell(row=r, column=c).fill = BG_FILL

    # Column widths for clean grid (#10 alignment)
    for c in range(1, 22):
        ws.column_dimensions[get_column_letter(c)].width = 9.5
    ws.column_dimensions["A"].width = 2  # left margin

    # ── HEADER BANNER (#2 theme) ──
    for c in range(1, 22):
        for r in [1, 2, 3]:
            ws.cell(row=r, column=c).fill = DARK_FILL

    set_cell(ws, 1, 2, "NHS England",
             Font("Calibri", bold=False, color=NHS_LIGHT, size=9), DARK_FILL,
             Alignment(horizontal="left"))
    set_cell(ws, 2, 2, "Referral to Treatment (RTT) Waiting Times",
             Font("Calibri", bold=True, color=WHITE, size=16), DARK_FILL,
             Alignment(horizontal="left"), merge_end_col=15)

    latest = nat_trend["Period"].iloc[-1]
    earliest = nat_trend["Period"].iloc[0]
    set_cell(ws, 3, 2,
             f"Monthly Incomplete Pathways  |  ICB Commissioner Level  |  {earliest} to {latest}",
             Font("Calibri", color=NHS_LIGHT, size=9), DARK_FILL,
             Alignment(horizontal="left"), merge_end_col=15)

    # ── KPI CARDS (#5 dynamic formulas, #11 readable, #12 K/M format) ──
    # KPIs use INDEX formulas referencing National_Trend so they update dynamically
    nat_last_row = len(nat_trend) + 1  # +1 for header
    nat_prev_row = nat_last_row - 1

    # National_Trend columns: A=Period, B=TF Code, C=TF, D=Pathways,
    # E=Within18, F=%18, G=Median, H=92nd, I=52+, J=65+, K=78+
    kpi_defs = [
        {
            "label": "Total Incomplete\nPathways",
            "formula": f"=National_Trend!D{nat_last_row}",
            "fmt": NUM_K,
            "change_formula": f"=National_Trend!D{nat_last_row}-National_Trend!D{nat_prev_row}",
            "change_fmt": NUM_K,
            "higher_bad": True,
        },
        {
            "label": "% Within\n18 Weeks",
            "formula": f"=National_Trend!F{nat_last_row}",
            "fmt": PCT_FMT,
            "change_formula": f"=National_Trend!F{nat_last_row}-National_Trend!F{nat_prev_row}",
            "change_fmt": "+0.0%;-0.0%",
            "higher_bad": False,
        },
        {
            "label": "Median Wait\n(Weeks)",
            "formula": f"=National_Trend!G{nat_last_row}",
            "fmt": DEC_FMT,
            "change_formula": f"=National_Trend!G{nat_last_row}-National_Trend!G{nat_prev_row}",
            "change_fmt": "+0.0;-0.0",
            "higher_bad": True,
        },
        {
            "label": "52+ Week\nWaiters",
            "formula": f"=National_Trend!I{nat_last_row}",
            "fmt": NUM_K,
            "change_formula": f"=National_Trend!I{nat_last_row}-National_Trend!I{nat_prev_row}",
            "change_fmt": NUM_K,
            "higher_bad": True,
        },
    ]

    kpi_start_row = 5
    for i, kpi in enumerate(kpi_defs):
        col_s = 2 + i * 5
        col_e = col_s + 3

        # White card background
        for r in range(kpi_start_row, kpi_start_row + 4):
            for c in range(col_s, col_e + 1):
                ws.cell(row=r, column=c).fill = WHITE_FILL
                ws.cell(row=r, column=c).border = Border(
                    left=Side(style="thin", color=NHS_PALE),
                    right=Side(style="thin", color=NHS_PALE),
                    top=Side(style="thin", color=NHS_PALE),
                    bottom=Side(style="thin", color=NHS_PALE),
                )

        # Top accent line
        for c in range(col_s, col_e + 1):
            ws.cell(row=kpi_start_row, column=c).border = Border(
                top=Side(style="medium", color=NHS_BLUE),
                left=Side(style="thin", color=NHS_PALE),
                right=Side(style="thin", color=NHS_PALE),
            )

        # Label (#6 consistent size)
        set_cell(ws, kpi_start_row, col_s, kpi["label"],
                 Font("Calibri", bold=True, color=NHS_GREY, size=9), WHITE_FILL,
                 Alignment(horizontal="center", vertical="center", wrap_text=True),
                 merge_end_col=col_e)

        # Value (dynamic formula #5)
        val_cell = set_cell(ws, kpi_start_row + 1, col_s, kpi["formula"],
                            Font("Calibri", bold=True, color=NHS_DARK, size=22),
                            WHITE_FILL, Alignment(horizontal="center"),
                            merge_end_col=col_e)
        val_cell.number_format = kpi["fmt"]

        # Change vs previous month (dynamic formula #5)
        chg_cell = set_cell(ws, kpi_start_row + 2, col_s, kpi["change_formula"],
                            Font("Calibri", color=NHS_GREY, size=9), WHITE_FILL,
                            Alignment(horizontal="center"),
                            merge_end_col=col_e)
        chg_cell.number_format = kpi["change_fmt"]

        # "vs prev month" label
        set_cell(ws, kpi_start_row + 3, col_s, "vs previous month",
                 Font("Calibri", italic=True, color=NHS_GREY, size=8), WHITE_FILL,
                 Alignment(horizontal="center"), merge_end_col=col_e)

    # Row heights for KPI section
    ws.row_dimensions[kpi_start_row].height = 30
    ws.row_dimensions[kpi_start_row + 1].height = 38
    ws.row_dimensions[kpi_start_row + 2].height = 18
    ws.row_dimensions[kpi_start_row + 3].height = 15

    # ── CHARTS (#7: exactly 5 charts, #8: varied types) ──
    # Layout: 2 columns, aligned (#10)
    # Row 1 of charts: B10 (left), L10 (right)
    # Row 2 of charts: B26 (left), L26 (right)
    # Row 3 of charts: B42 (centre/full)

    ws_nat = wb["National_Trend"]
    nrows_nat = len(nat_trend)
    ws_icb = wb["ICB_Summary"]
    ws_spec = wb["Specialty_Top5"]
    nrows_spec = len(spec_top5)
    ws_tb = wb["Top5_Bottom5"]

    # Chart 1 (LINE) – Total pathways monthly trend (#15 clear time label)
    add_line_chart(
        ws, ws_nat,
        f"Monthly Total Incomplete Pathways ({earliest} to {latest})",
        "Total number of incomplete pathways",
        nrows_nat, "B10", color=NHS_BLUE, y_fmt=NUM_K,
    )

    # Chart 2 (PIE) – Top 5 specialties + Other (#14 limited to 6 slices)
    add_pie_chart(
        ws, ws_spec,
        f"Share of Waiting List by Specialty ({latest})",
        "Specialty", "Incomplete Pathways",
        nrows_spec, "L10",
    )

    # Chart 3 (DUAL LINE) – 52+ vs 65+ week waiters comparison (#8 variety)
    add_dual_line_chart(
        ws, ws_nat,
        f"Long Waiters: 52+ Weeks vs 65+ Weeks ({earliest} to {latest})",
        "Total 52 plus weeks",
        "Total 65 plus weeks",
        nrows_nat, "B26",
    )

    # Chart 4 (HORIZONTAL BAR) – Top 5 best ICBs by % within 18wks (#13 top 5)
    # ICB_Summary is sorted by % within 18 weeks desc, so first 5 rows = top 5
    chart4 = BarChart()
    chart4.type = "bar"
    chart4.title = f"Top 5 ICBs: Highest % Within 18 Weeks ({latest})"
    chart4.style = 2
    chart4.width = 22
    chart4.height = 13
    data4 = Reference(ws_icb, min_col=8, min_row=1, max_row=6)  # % within 18wks, top 5
    cats4 = Reference(ws_icb, min_col=3, min_row=2, max_row=6)  # ICB Name, top 5
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.series[0].graphicalProperties.solidFill = NHS_GREEN
    chart4.x_axis.numFmt = PCT_FMT
    chart4.legend = None
    ws.add_chart(chart4, "L26")

    # Chart 5 (LINE) – 52+ week waiters trend
    add_line_chart(
        ws, ws_nat,
        f"52+ Week Waiters Trend ({earliest} to {latest})",
        "Total 52 plus weeks",
        nrows_nat, "B42", color=NHS_RED, y_fmt=NUM_K,
    )

    # ── BOTTOM 5 TABLE (next to chart 5, #13) ──
    table_start_row = 42
    set_cell(ws, table_start_row, 14,
             f"Bottom 5 ICBs: Lowest % Within 18 Weeks ({latest})",
             Font("Calibri", bold=True, color=WHITE, size=11),
             PatternFill(start_color=NHS_RED, end_color=NHS_RED, fill_type="solid"),
             Alignment(horizontal="center"), merge_end_col=20)
    for c in range(14, 21):
        ws.cell(row=table_start_row, column=c).fill = PatternFill(
            start_color=NHS_RED, end_color=NHS_RED, fill_type="solid")
        ws.cell(row=table_start_row, column=c).font = Font(
            "Calibri", bold=True, color=WHITE, size=11)

    # Table headers
    tb_headers = ["Rank", "ICB", "% Within 18 Wks", "Pathways", "Median Wait", "52+ Waiters"]
    for ci, h in enumerate(tb_headers):
        cell = ws.cell(row=table_start_row + 1, column=14 + ci, value=h)
        cell.font = Font("Calibri", bold=True, color=WHITE, size=9)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for ri, (_, row_data) in enumerate(worst5.iterrows()):
        for ci, col_name in enumerate(worst5.columns):
            val = row_data[col_name]
            cell = ws.cell(row=table_start_row + 2 + ri, column=14 + ci + 1, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = WHITE_FILL
            if col_name == "% Within 18 Wks":
                cell.number_format = PCT_FMT
            elif isinstance(val, (int, float)):
                cell.number_format = NUM_K
        # Rank column
        rank_cell = ws.cell(row=table_start_row + 2 + ri, column=14, value=ri + 1)
        rank_cell.border = THIN_BORDER
        rank_cell.alignment = Alignment(horizontal="center")
        rank_cell.fill = WHITE_FILL

    # ── INSIGHTS SECTION ──
    insight_row = 56
    set_cell(ws, insight_row, 2, "Key Insights",
             Font("Calibri", bold=True, color=NHS_DARK, size=13), BG_FILL,
             Alignment(horizontal="left"), merge_end_col=6)

    # Blue accent line
    for c in range(2, 21):
        ws.cell(row=insight_row + 1, column=c).border = Border(
            top=Side(style="medium", color=NHS_BLUE))

    # Compute insight values
    first_row = nat_trend.iloc[0]
    last_row = nat_trend.iloc[-1]
    first_pathways = safe_int(first_row["Total number of incomplete pathways"])
    last_pathways = safe_int(last_row["Total number of incomplete pathways"])
    path_change_pct = abs((last_pathways - first_pathways) / first_pathways * 100)

    first_52 = safe_int(first_row["Total 52 plus weeks"])
    last_52 = safe_int(last_row["Total 52 plus weeks"])
    wk52_change_pct = abs((last_52 - first_52) / first_52 * 100)

    first_pct18 = safe_float(first_row["% within 18 weeks"])
    last_pct18 = safe_float(last_row["% within 18 weeks"])

    best_icb = icb_summary.iloc[0]
    worst_icb = icb_summary.iloc[-1]

    best_name = str(best_icb["ICB Name"]).replace("NHS ", "").replace(" INTEGRATED CARE BOARD", "")
    worst_name = str(worst_icb["ICB Name"]).replace("NHS ", "").replace(" INTEGRATED CARE BOARD", "")

    insights = [
        f"Total incomplete pathways fell {path_change_pct:.0f}% from {fmt_k(first_pathways)} to {fmt_k(last_pathways)} over the 42-month period.",
        f"52+ week waiters dropped {wk52_change_pct:.0f}% from {fmt_k(first_52)} to {fmt_k(last_52)}, a significant reduction in long waits.",
        f"% treated within 18 weeks improved from {first_pct18:.1%} to {last_pct18:.1%}, though still below the 92% NHS constitutional target.",
        f"Best performing ICB: {best_name} ({safe_float(best_icb['% within 18 weeks']):.1%} within 18 weeks).",
        f"Most challenged ICB: {worst_name} ({safe_float(worst_icb['% within 18 weeks']):.1%} within 18 weeks).",
    ]

    for j, text in enumerate(insights):
        bullet_col = 2
        set_cell(ws, insight_row + 2 + j, bullet_col, text,
                 Font("Calibri", color=BLACK, size=10), BG_FILL,
                 Alignment(horizontal="left", wrap_text=True, indent=1),
                 merge_end_col=20)
        ws.row_dimensions[insight_row + 2 + j].height = 20

    # ── FOOTER ──
    footer_row = insight_row + 2 + len(insights) + 1
    set_cell(ws, footer_row, 2,
             "Source: NHS England RTT Statistics  |  england.nhs.uk/statistics  |  Data updated monthly",
             Font("Calibri", italic=True, color=NHS_GREY, size=8), BG_FILL,
             Alignment(horizontal="left"), merge_end_col=15)
    set_cell(ws, footer_row + 1, 2,
             "Tip: Use the Data sheet filters or add Slicers (Insert > Slicer) for interactive filtering.",
             Font("Calibri", italic=True, color=NHS_BLUE, size=8), BG_FILL,
             Alignment(horizontal="left"), merge_end_col=15)

    # (#1) Grid lines off, (#3) no field buttons (not applicable — no PivotTables)
    ws.sheet_view.showGridLines = False


# ── Main builder ──────────────────────────────────────────────────────


def main():
    print("Loading data...")
    nat_df = pd.read_csv(NAT_CSV)
    icb_df = pd.read_csv(ICB_CSV)
    for col in ["Average (median) waiting time (in weeks)",
                 "92nd percentile waiting time (in weeks)"]:
        nat_df[col] = pd.to_numeric(nat_df[col], errors="coerce")
        icb_df[col] = pd.to_numeric(icb_df[col], errors="coerce")
    latest = nat_df["Period"].max()
    print(f"Latest period: {latest}")
    print(f"National: {len(nat_df):,} rows | ICB: {len(icb_df):,} rows")

    print("Building summary tables...")
    nat_trend = build_national_trend(nat_df)
    icb_summary = build_icb_summary(icb_df, latest)
    spec_top5 = build_specialty_top5(nat_df, latest)
    best5, worst5 = build_top5_bottom5(icb_df, latest)
    yoy = build_yoy_change(nat_df)

    pct_cols = ["% within 18 weeks"]
    wb = Workbook()

    # ── Sheet: Data (ICB-level as Excel Table for slicer support #9) ──
    ws_data = wb.active
    ws_data.title = "Data"
    data_trimmed = icb_df[icb_df["Treatment Function"] != "Total"][ICB_KEEP].copy()
    data_trimmed.columns = SHORT_HEADERS
    write_df(ws_data, data_trimmed, ["Pct_Within_18wks"])

    # (#12) Format large numbers as K in the data sheet
    num_cols_data = ["Incomplete_Pathways", "Within_18wks", "Waiters_52plus",
                     "Waiters_65plus", "Waiters_78plus"]
    for col_name in num_cols_data:
        if col_name in SHORT_HEADERS:
            ci = SHORT_HEADERS.index(col_name) + 1
            col_l = get_column_letter(ci)
            for r in range(2, len(data_trimmed) + 2):
                ws_data.cell(row=r, column=ci).number_format = NUM_K

    # Create Excel Table (#9 slicer-ready)
    last_col_letter = get_column_letter(len(SHORT_HEADERS))
    table_ref = f"A1:{last_col_letter}{len(data_trimmed) + 1}"
    tab = Table(displayName="RTTData", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws_data.add_table(tab)
    ws_data.freeze_panes = "A2"

    # Data bar on % within 18 weeks
    pct_ci = SHORT_HEADERS.index("Pct_Within_18wks") + 1
    pct_l = get_column_letter(pct_ci)
    ws_data.conditional_formatting.add(
        f"{pct_l}2:{pct_l}{len(data_trimmed)+1}",
        DataBarRule(start_type="min", end_type="max", color=NHS_GREEN),
    )
    print(f"  Data: {len(data_trimmed):,} rows (Excel Table 'RTTData')")

    # ── Sheet: National_Trend ──
    ws_nat = wb.create_sheet("National_Trend")
    write_df(ws_nat, nat_trend, pct_cols)
    ws_nat.freeze_panes = "A2"
    print(f"  National_Trend: {len(nat_trend)} rows")

    # ── Sheet: ICB_Summary ──
    ws_icb = wb.create_sheet("ICB_Summary")
    write_df(ws_icb, icb_summary, pct_cols)
    ws_icb.auto_filter.ref = f"A1:{get_column_letter(len(ICB_KEEP))}1"
    ws_icb.freeze_panes = "A2"
    icb_pct_col = get_column_letter(ICB_KEEP.index("% within 18 weeks") + 1)
    ws_icb.conditional_formatting.add(
        f"{icb_pct_col}2:{icb_pct_col}{len(icb_summary)+1}",
        ColorScaleRule(
            start_type="min", start_color=NHS_RED,
            mid_type="percentile", mid_value=50, mid_color=NHS_AMBER,
            end_type="max", end_color=NHS_GREEN,
        ),
    )
    print(f"  ICB_Summary: {len(icb_summary)} rows")

    # ── Sheet: Specialty_Top5 ──
    ws_spec = wb.create_sheet("Specialty_Top5")
    write_df(ws_spec, spec_top5)
    # Format pathways as K
    for r in range(2, len(spec_top5) + 2):
        ws_spec.cell(row=r, column=2).number_format = NUM_K
    ws_spec.freeze_panes = "A2"
    print(f"  Specialty_Top5: {len(spec_top5)} rows")

    # ── Sheet: Top5_Bottom5 ──
    ws_tb = wb.create_sheet("Top5_Bottom5")
    # Top 5 banner
    for c in range(1, 6):
        ws_tb.cell(row=1, column=c).fill = PatternFill(
            start_color=NHS_GREEN, end_color=NHS_GREEN, fill_type="solid")
        ws_tb.cell(row=1, column=c).font = Font("Calibri", bold=True, color=WHITE, size=12)
    set_cell(ws_tb, 1, 1, f"Top 5 ICBs — Highest % Within 18 Weeks ({latest})",
             Font("Calibri", bold=True, color=WHITE, size=12), merge_end_col=5)
    write_df(ws_tb, best5, ["% Within 18 Wks"], start_row=2)
    for r in range(3, 8):
        ws_tb.cell(row=r, column=3).number_format = NUM_K
        ws_tb.cell(row=r, column=5).number_format = NUM_K

    gap = len(best5) + 4
    for c in range(1, 6):
        ws_tb.cell(row=gap, column=c).fill = PatternFill(
            start_color=NHS_RED, end_color=NHS_RED, fill_type="solid")
        ws_tb.cell(row=gap, column=c).font = Font("Calibri", bold=True, color=WHITE, size=12)
    set_cell(ws_tb, gap, 1, f"Bottom 5 ICBs — Lowest % Within 18 Weeks ({latest})",
             Font("Calibri", bold=True, color=WHITE, size=12), merge_end_col=5)
    write_df(ws_tb, worst5, ["% Within 18 Wks"], start_row=gap + 1)
    for r in range(gap + 2, gap + 7):
        ws_tb.cell(row=r, column=3).number_format = NUM_K
        ws_tb.cell(row=r, column=5).number_format = NUM_K

    auto_width(ws_tb)
    print(f"  Top5_Bottom5: {len(best5)} + {len(worst5)} rows")

    # ── Sheet: YoY_Change ──
    ws_yoy = wb.create_sheet("YoY_Change")
    yoy_pct = ["% Within 18 Wks", "% Within 18 Wks YoY"]
    write_df(ws_yoy, yoy, yoy_pct)
    ws_yoy.freeze_panes = "A2"
    for col_name in [c for c in yoy.columns if "YoY" in c]:
        ci = list(yoy.columns).index(col_name) + 1
        cl = get_column_letter(ci)
        rng = f"{cl}2:{cl}{len(yoy)+1}"
        ws_yoy.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0"],
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            font=Font(color=NHS_GREEN),
        ))
        ws_yoy.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThan", formula=["0"],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            font=Font(color=NHS_RED),
        ))
    print(f"  YoY_Change: {len(yoy)} rows")

    # ── Sheet: Lookups ──
    ws_look = wb.create_sheet("Lookups")
    icb_names = sorted(icb_df["ICB Name"].unique())
    ws_look.cell(row=1, column=1, value="ICB Name").font = HEADER_FONT
    ws_look.cell(row=1, column=1).fill = HEADER_FILL
    for i, name in enumerate(icb_names, 2):
        ws_look.cell(row=i, column=1, value=name)

    periods = sorted(nat_df["Period"].unique())
    ws_look.cell(row=1, column=3, value="Period").font = HEADER_FONT
    ws_look.cell(row=1, column=3).fill = HEADER_FILL
    for i, p in enumerate(periods, 2):
        ws_look.cell(row=i, column=3, value=p)

    tfs = sorted(nat_df[nat_df["Treatment Function"] != "Total"]["Treatment Function"].unique())
    ws_look.cell(row=1, column=5, value="Treatment Function").font = HEADER_FONT
    ws_look.cell(row=1, column=5).fill = HEADER_FILL
    for i, t in enumerate(tfs, 2):
        ws_look.cell(row=i, column=5, value=t)
    auto_width(ws_look)
    print(f"  Lookups: {len(icb_names)} ICBs, {len(periods)} periods, {len(tfs)} specialties")

    # ── Sheet: Dashboard ──
    print("Building dashboard...")
    build_dashboard(wb, nat_trend, icb_summary, spec_top5, best5, worst5)

    # ── Reorder sheets ──
    order = ["Dashboard", "Data", "National_Trend", "ICB_Summary",
             "Specialty_Top5", "Top5_Bottom5", "YoY_Change", "Lookups"]
    for i, name in enumerate(order):
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")


if __name__ == "__main__":
    main()
